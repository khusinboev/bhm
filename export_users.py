"""Foydalanuvchilar bazasini eksport qilish (viloyat + telefon).

Ma'lumot har bir bot nusxasining o'z Postgres bazasidagi `accounts`
jadvalida turadi (bhm, bhm2, bhm3). Skript uchalasini birlashtirib
bitta Excel fayl tayyorlaydi.

Ishga tushirish:
    <venv>/bin/python export_users.py                 # Excel yaratish
    <venv>/bin/python export_users.py --csv           # CSV ham
    <venv>/bin/python export_users.py --send 12345678 # tayyor faylni yuborish
    <venv>/bin/python export_users.py --all           # ro'yxatdan o'tmaganlar ham
"""

import asyncio
import csv
import io
import os
import sys
from datetime import datetime

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import DB_CONFIG, BOT_TOKEN  # noqa: E402

# Bot nusxalari va ularning bazalari
BAZALAR = [("Asosiy bot", "bhm"), ("Klon-1", "bhm2"), ("Klon-2", "bhm3")]

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")


def _fetch(dbname: str, only_registered: bool) -> list[tuple]:
    cfg = {**DB_CONFIG, "dbname": dbname}
    shart = "WHERE viloyat IS NOT NULL OR phone IS NOT NULL" if only_registered else ""
    try:
        conn = psycopg2.connect(**cfg)
    except psycopg2.Error as e:
        print(f"  ⚠️  {dbname} bazasiga ulanib bo'lmadi: {str(e).strip()}")
        return []
    try:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT user_id, viloyat, phone, date
                FROM public.accounts {shart}
                ORDER BY date DESC NULLS LAST, user_id
            """)
            return cur.fetchall()
    finally:
        conn.close()


def _style(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def build_excel(rows: list[tuple], stats: list[tuple]) -> bytes:
    wb = Workbook()

    # --- Umumiy ---
    ws = wb.active
    ws.title = "Umumiy"
    ws.append(["Bot", "Jami foydalanuvchi", "Viloyat bergan", "Telefon bergan", "To'liq"])
    _style(ws, 5)
    for nom, jami, v, p, toliq in stats:
        ws.append([nom, jami, v, p, toliq])
    ws.append(["JAMI", sum(s[1] for s in stats), sum(s[2] for s in stats),
               sum(s[3] for s in stats), sum(s[4] for s in stats)])
    for c in range(1, 6):
        ws.cell(ws.max_row, c).font = Font(bold=True)
    for r in range(2, ws.max_row + 1):
        for c in range(2, 6):
            ws.cell(r, c).number_format = "#,##0"
    for i, w in enumerate([16, 20, 16, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Foydalanuvchilar ---
    ws2 = wb.create_sheet("Foydalanuvchilar")
    ws2.append(["Bot", "Telegram ID", "Viloyat", "Telefon", "Sana"])
    _style(ws2, 5)
    for bot_nom, uid, viloyat, phone, sana in rows:
        ws2.append([bot_nom, uid, viloyat or "", phone or "",
                    sana.strftime("%Y-%m-%d %H:%M") if sana else ""])
    for row in ws2.iter_rows(min_row=2, min_col=2, max_col=2):
        for c in row:
            c.number_format = "0"
    # Telefon matn sifatida qolsin (boshidagi + yo'qolmasin)
    for row in ws2.iter_rows(min_row=2, min_col=4, max_col=4):
        for c in row:
            c.alignment = Alignment(horizontal="left")
    for i, w in enumerate([14, 16, 20, 18, 18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.auto_filter.ref = ws2.dimensions

    # --- Viloyatlar kesimi ---
    ws3 = wb.create_sheet("Viloyatlar")
    ws3.append(["Viloyat", "Soni", "Ulush"])
    _style(ws3, 3)
    hisob: dict[str, int] = {}
    for _, _, viloyat, _, _ in rows:
        if viloyat:
            hisob[viloyat] = hisob.get(viloyat, 0) + 1
    jami_v = sum(hisob.values()) or 1
    for v, n in sorted(hisob.items(), key=lambda x: -x[1]):
        ws3.append([v, n, n / jami_v])
        ws3.cell(ws3.max_row, 2).number_format = "#,##0"
        ws3.cell(ws3.max_row, 3).number_format = "0.0%"
    for i, w in enumerate([24, 12, 10], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv(rows: list[tuple]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["bot", "telegram_id", "viloyat", "telefon", "sana"])
    for bot_nom, uid, viloyat, phone, sana in rows:
        w.writerow([bot_nom, uid, viloyat or "", phone or "",
                    sana.strftime("%Y-%m-%d %H:%M") if sana else ""])
    return buf.getvalue().encode("utf-8-sig")  # Excel uchun BOM


async def send_file(chat_id: int, data: bytes, filename: str, caption: str) -> None:
    import aiohttp
    api = f"https://api.telegram.org/bot{BOT_TOKEN}"
    form = aiohttp.FormData()
    form.add_field("chat_id", str(chat_id))
    form.add_field("caption", caption)
    form.add_field("parse_mode", "HTML")
    form.add_field("document", data, filename=filename,
                   content_type="application/octet-stream")
    async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=120)) as s:
        async with s.post(f"{api}/sendDocument", data=form) as r:
            d = await r.json()
    print("Yuborildi ✅" if d.get("ok") else f"Yuborib bo'lmadi: {d.get('description')}")


def main() -> None:
    args = sys.argv[1:]
    only_registered = "--all" not in args

    rows, stats = [], []
    for nom, dbname in BAZALAR:
        barcha = _fetch(dbname, only_registered=False)
        if not barcha:
            continue
        v = sum(1 for r in barcha if r[1])
        p = sum(1 for r in barcha if r[2])
        toliq = sum(1 for r in barcha if r[1] and r[2])
        stats.append((nom, len(barcha), v, p, toliq))
        for r in barcha:
            if only_registered and not (r[1] or r[2]):
                continue
            rows.append((nom, r[0], r[1], r[2], r[3]))
        print(f"{nom} ({dbname}): jami {len(barcha):,}, viloyat {v:,}, telefon {p:,}")

    if not stats:
        print("Hech qanday bazaga ulanib bo'lmadi.")
        return

    print(f"\nEksportga tushadigan yozuvlar: {len(rows):,}"
          f"{'' if only_registered else ' (barcha userlar)'}")

    sana = datetime.now().strftime("%Y%m%d_%H%M")
    xlsx = build_excel(rows, stats)
    xlsx_name = f"foydalanuvchilar_{sana}.xlsx"
    with open(xlsx_name, "wb") as f:
        f.write(xlsx)
    print(f"Excel: {xlsx_name} ({len(xlsx):,} bayt)")

    if "--csv" in args:
        csv_data = build_csv(rows)
        csv_name = f"foydalanuvchilar_{sana}.csv"
        with open(csv_name, "wb") as f:
            f.write(csv_data)
        print(f"CSV: {csv_name} ({len(csv_data):,} bayt)")

    if "--send" in args:
        i = args.index("--send")
        if i + 1 >= len(args):
            print("--send uchun chat_id kerak")
            return
        chat_id = int(args[i + 1])
        caption = (f"👥 <b>Foydalanuvchilar bazasi</b>\n"
                   f"{len(rows):,} ta yozuv | "
                   f"{datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
                   f"3 varaq: Umumiy, Foydalanuvchilar, Viloyatlar")
        asyncio.run(send_file(chat_id, xlsx, xlsx_name, caption))


if __name__ == "__main__":
    main()
